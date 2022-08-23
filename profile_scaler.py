import openpyxl as xl
from pathlib import Path
import pandas


class ProfileScaler:
    input_wbk: xl.Workbook = None

    def __init__(self):
        # self.input_wbk=xl.load_workbook('Prod_Prof_scale.xlsx')
        pass

    def __del__(self):
        self.input_wbk.close()

    def set_workbook(self, wbk_path: Path) -> None:
        print("\nLoading workbook...\n")
        self.input_wbk = xl.load_workbook(wbk_path)

    def get_required_columns(self, input_df: pandas.DataFrame):
        # df=pd.read_excel(workbook,sheet_name=worksheet,header=[0,1],index_col=0,parse_dates=True,engine='openpyxl')
        columns_to_delete = []
        for tup in input_df.columns.values:
            if "CUM" not in tup[1]:
                columns_to_delete.append(tup)
        for col in columns_to_delete:
            input_df.pop(col)
        return input_df

    def check_input_sheet_available(self) -> bool:
        if "Profile_Scaling_Input" in self.input_wbk.sheetnames:
            return True
        return False

    def get_scaling_parameters(self):
        parameter_table = [
            row for row in self.input_wbk["Profile_Scaling_Input"].values
        ]
        output_params=[]
        for i in range(1,len(parameter_table)):
            item=ScalingParameter(parameter_table[i])
            output_params.append(item)
        return output_params

    def scale_cums_to_target_from_date(self, input_arr, target_cum, from_date=None):
        scaling_factor = target_cum / input_arr[-1][1]
        if from_date is None:
            from_date = input_arr[0][0]
        output_cums = [
            cum * scaling_factor if date >= from_date else cum
            for date, cum in input_arr
        ]
        return output_cums

class ScalingParameter:
    entity_type=None
    entity_name=None
    property_name=None
    target_cum=None
    scaling_factor=None
    start_date=None
    other_properties=None
    
    def __init__(self,input_list: list):
        # for item in input_list:
        self.entity_type=input_list[0]
        self.entity_name=input_list[1]
        self.property_name=input_list[2]
        self.target_cum=input_list[3]
        self.scaling_factor=input_list[4]
        self.start_date=input_list[5]
        self.other_properties=input_list[6]
   
    def get_other_properties_list(self)->list:
        if self.other_properties is None:
            return None
        return self.other_properties.split(',')



