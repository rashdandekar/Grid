# import openpyxl as xl
from dataclasses import replace
import openpyxl as xl
import pandas as pd
import read_xl_profile as rxp
import datetime
import numpy as np

# import profile_scaler
from pathlib import Path


def get_data_for_scaling() -> dict:
    entity_types = ["USERDF", "SOURCE", "SEP", "TANK", "JOINT", "WELL", "INLGEN"]
    # entity_types = ["SEP"]
    df_dict = {}
    wbk = xl.load_workbook("Prod_Prof_scale.xlsx")
    for entity_type in entity_types:
        print(f"Starting to process {entity_type} sheet\n")
        input_df = pd.read_excel(
            wbk,
            sheet_name=entity_type,
            header=[0, 1],
            index_col=0,
            engine="openpyxl",
        )
        input_dates = input_df.index.to_numpy()
        input_df.index = input_dates

        input_df = rxp.get_required_columns(input_df)
        if not input_df.empty:
            df_dict[entity_type] = input_df
    return df_dict


def get_rates_from_cum_series(series: pd.Series):
    cums = list(series.to_numpy())
    dates = series.index.to_numpy()
    days = [(date - dates[0]) / np.timedelta64(1, "D") for date in dates]
    rates = []
    for i in range(len(cums) - 1):
        rates.append((cums[i + 1] - cums[i]) / (days[i + 1] - days[i]))
    rates.append(0)
    # rate_series=pd.Series(rates)
    # rate_series.index=series.index
    output_df = pd.DataFrame(pd.concat([pd.Series(cums), pd.Series(rates)], axis=1))
    output_df.index = series.index
    return output_df


def main() -> None:
    # data=get_data_for_scaling()
    df_dict = {}
    ps = ProfileScaler()
    ps.set_workbook(Path("Prod_Prof_scale.xlsx"))
    if ps.check_input_sheet_available():
        input_paramaters = ps.get_scaling_parameters()
        data = get_data_for_scaling()
        for input_parameter in input_paramaters:
            if input_parameter.entity_type in list(data.keys()):
                df = data[input_parameter.entity_type]
                df.index = [
                    datetime.datetime.fromisoformat(x[0 : len(x) - 1])
                    for x in df.index.to_numpy()
                ]
                df_to_scale = df[
                    (input_parameter.entity_name, input_parameter.property_name)
                ]
                scaled_cum = pd.Series(
                    ps.scale_cums_to_target_from_date(
                        list(zip(df_to_scale.index.to_numpy(), df_to_scale.to_numpy())),
                        input_parameter.target_cum,
                        input_parameter.start_date,
                    )
                )
                scaled_cum.index = df_to_scale.index
                # cols=df.columns.values
                new_header = [
                    (input_parameter.entity_name, input_parameter.property_name),
                    (
                        input_parameter.entity_name,
                        "AVG_"
                        + input_parameter.property_name.replace("CUM", "")
                        + "_RATE"
                    ),
                    (
                        input_parameter.entity_name,
                        "SCALED_" + input_parameter.property_name
                    ),
                    (
                        input_parameter.entity_name,
                        "SCALED_"
                        + "AVG_"
                        + input_parameter.property_name.replace("CUM", "")
                        + "_RATE"
                    ),
                ]
                original_cums_with_rates = get_rates_from_cum_series(df_to_scale)
                new_cums_with_rates = get_rates_from_cum_series(scaled_cum)
                new_df = pd.concat([original_cums_with_rates, new_cums_with_rates], axis=1)
                new_df.columns = pd.MultiIndex.from_tuples(new_header)
                df_dict[
                    "SCALED_"
                    + input_parameter.entity_name
                    + "_"
                    + input_parameter.property_name
                ] = new_df

    if len(df_dict) > 0:
        print(df_dict.keys())


class ProfileScaler:
    input_wbk: xl.Workbook = None

    def __init__(self):
        # self.input_wbk=xl.load_workbook('Prod_Prof_scale.xlsx')
        pass

    def __del__(self):
        self.input_wbk.close()

    def set_workbook(self, wbk_path: Path) -> None:
        print("\nLoading workbook...\n")
        self.input_wbk = xl.load_workbook(wbk_path)

    def get_required_columns(self, input_df: pd.DataFrame):
        # df=pd.read_excel(workbook,sheet_name=worksheet,header=[0,1],index_col=0,parse_dates=True,engine='openpyxl')
        columns_to_delete = []
        for tup in input_df.columns.values:
            if "CUM" not in tup[1]:
                columns_to_delete.append(tup)
        for col in columns_to_delete:
            input_df.pop(col)
        return input_df

    def check_input_sheet_available(self) -> bool:
        if "Profile_Scaling_Input" in self.input_wbk.sheetnames:
            return True
        return False

    def get_scaling_parameters(self):
        parameter_table = [
            row for row in self.input_wbk["Profile_Scaling_Input"].values
        ]
        output_params = []
        for i in range(1, len(parameter_table)):
            item = ScalingParameter(parameter_table[i])
            output_params.append(item)
        return output_params

    def scale_cums_to_target_from_date(self, input_arr, target_cum, from_date=None):
        scaling_factor = target_cum / input_arr[-1][1]
        if from_date is None:
            from_date = input_arr[0][0]
        output_cums = [
            cum * scaling_factor if date >= from_date else cum
            for date, cum in input_arr
        ]
        return output_cums


class ScalingParameter:
    entity_type = None
    entity_name = None
    property_name = None
    target_cum = None
    scaling_factor = None
    start_date = None
    other_properties = None

    def __init__(self, input_list: list):
        # for item in input_list:
        self.entity_type = input_list[0]
        self.entity_name = input_list[1]
        self.property_name = input_list[2]
        self.target_cum = input_list[3]
        self.scaling_factor = input_list[4]
        self.start_date = input_list[5]
        self.other_properties = input_list[6]

    def get_other_properties_list(self) -> list:
        if self.other_properties is None:
            return None
        return self.other_properties.split(",")


if __name__ == "__main__":
    main()
